# -*- coding: utf-8 -*-
"""Merge idiom phrase sources into data/passage_phrases.json + compact JS lookup."""
import json
import re
from pathlib import Path

import openpyxl

try:
    import fitz
except ImportError:
    fitz = None

SRC = Path(r"g:\트리거 앱\시중 단어 교재\숙어모음")
OUT_JSON = Path(__file__).resolve().parent.parent / "data" / "passage_phrases.json"
OUT_JS = Path(__file__).resolve().parent.parent / "data" / "passage_phrases_lookup.js"

HAS_HANGUL = re.compile(r"[\uac00-\ud7a3]")
HAS_LATIN = re.compile(r"[a-z]", re.I)
SKIP_FILE_HINTS = ("정오표", "EBS수능특강시리즈정오")


def norm_phrase(s: str) -> str:
    s = str(s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w\s'.,\-–—/()]", "", s)
    return s.strip()


def norm_meaning(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.replace(";", ",").strip(" ,")


def add(index: dict, phrase: str, meaning: str, source: str) -> bool:
    p = norm_phrase(phrase)
    m = norm_meaning(meaning)
    if not p or len(p) < 3 or not m or not HAS_HANGUL.search(m):
        return False
    if not HAS_LATIN.search(p):
        return False
    if " " not in p and "-" not in p:
        return False
    if len(p) > 90:
        return False
    if any(x in p for x in ("http", "www", "day ", "lesson")):
        return False
    if p not in index:
        index[p] = {"mean": m, "sources": [source]}
        return True
    if source not in index[p]["sources"]:
        index[p]["sources"].append(source)
    return False


def parse_middle_1000(path: Path, index: dict) -> int:
    n = 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue
        eng = row[2] if len(row) > 2 else None
        kor = row[3] if len(row) > 3 else None
        if eng and kor and add(index, str(eng), str(kor), "중학숙어1000"):
            n += 1
    wb.close()
    return n


def parse_suneung_gicho(path: Path, index: dict) -> int:
    n = 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue
        eng = row[1]
        kor = row[2]
        if eng and kor and add(index, str(eng), str(kor), "수능기출"):
            n += 1
    wb.close()
    return n


def parse_pdf_lines(path: Path, tag: str, index: dict) -> int:
    if not fitz:
        return 0
    n = 0
    doc = fitz.open(path)
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    for line in text.splitlines():
        line = line.strip()
        if len(line) < 6:
            continue
        for sep in ["\t", "|", "  ", " - ", " – "]:
            if sep in line:
                a, b = line.split(sep, 1)
                if HAS_LATIN.search(a) and HAS_HANGUL.search(b) and add(index, a, b, tag):
                    n += 1
                break
    return n


def main() -> None:
    index: dict = {}
    for path in sorted(SRC.iterdir()):
        if not path.is_file():
            continue
        name = path.name
        if any(h in name for h in SKIP_FILE_HINTS):
            print(f"SKIP {name}")
            continue
        try:
            if "1000" in name and path.suffix.lower() == ".xlsx":
                c = parse_middle_1000(path, index)
            elif "기출" in name and path.suffix.lower() == ".xlsx":
                c = parse_suneung_gicho(path, index)
            elif path.suffix.lower() == ".xlsx":
                c = parse_suneung_gicho(path, index)
            elif path.suffix.lower() == ".pdf":
                c = parse_pdf_lines(path, path.stem[:16], index)
            else:
                continue
            print(f"OK {name}: +{c} total={len(index)}")
        except Exception as e:
            print(f"ERR {name}: {e}")

    phrases = sorted(index.keys(), key=lambda x: (-len(x.split()), -len(x), x))
    payload = {
        "meta": {"count": len(phrases), "match_order": "longest_first"},
        "phrases": {p: {"mean": index[p]["mean"], "sources": index[p]["sources"]} for p in phrases},
    }
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "/* auto-generated; do not edit */",
        "window.PASSAGE_PHRASE_MEANINGS=(function(){var m=Object.create(null);",
    ]
    for p in phrases:
        lines.append(f"m[{json.dumps(p)}]={json.dumps(index[p]['mean'])};")
    lines.append("return m;})();")
    lines.append("window.PASSAGE_PHRASE_LIST=" + json.dumps(phrases, ensure_ascii=False) + ";")
    OUT_JS.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {len(phrases)} -> {OUT_JS} ({OUT_JS.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
