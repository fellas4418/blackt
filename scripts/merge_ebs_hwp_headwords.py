# -*- coding: utf-8 -*-
"""Merge EBS 수능특강 vocab (xlsx preferred, pdf supplement, hwp fallback). Does NOT modify worddata.js."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from extract_hwp_text import extract_paragraphs  # noqa: E402
from merge_market_headwords import add, norm  # noqa: E402

XLSX_PATH = Path(r"g:\트리거 앱\시중 단어 교재\2027 수능특강 영어 단어.xlsx")
PDF_PATH = Path(
    r"g:\트리거 앱\시중 단어 교재\EBS_2026학년도_수능특강_영어영역_영어_영단어·숙어(단어장).pdf"
)
HWP_PATH = Path(r"g:\트리거 앱\시중 단어 교재\2027 수능특강 영어 단어 (전체).hwp")
OUT = ROOT.parent / "data" / "merged_market_headwords.json"
OUT_JS = ROOT.parent / "data" / "market_headwords_lookup.js"
PHRASE_JSON = ROOT.parent / "data" / "passage_phrases.json"
PHRASE_JS = ROOT.parent / "data" / "passage_phrases_lookup.js"

SOURCE_XLSX = "EBS2027수능특강XLSX"
SOURCE_PDF = "EBS2026수능특강PDF"
SOURCE_HWP = "EBS2027수능특강HWP"

MARK = "\u2245"
HAS_HANGUL = re.compile(r"[\uac00-\ud7a3]")
HAS_LATIN = re.compile(r"[a-zA-Z]")
ONLY_NUM = re.compile(r"^\d{1,3}$")
LESSON_SUFFIX = re.compile(r"\d+강\d+번\s*$")
JUNK_LINE = re.compile(
    r"^(Date|Name|English|Korean|점수|氠|捤|潴|Date|상담문의|전강|콕!|수능특강)",
    re.I,
)
ENG_LINE = re.compile(r"^[a-zA-Z][a-zA-Z0-9\s'.,\-/~()\u2019]*$")


def eng_key(phrase: str) -> str:
    return re.sub(r"\s+", " ", phrase.strip().lower().replace("\u2019", "'"))


def tokens_from_phrase(phrase: str) -> list[str]:
    phrase = phrase.strip().replace("\u2019", "'")
    words: list[str] = []
    for part in re.split(r"[\s/]+", phrase):
        part = part.strip(".,;:")
        if part and norm(part):
            words.append(norm(part))
    return words


def parse_xlsx_pairs(path: Path) -> list[tuple[str, str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "2027년수능특강영단어"
    if sheet not in wb.sheetnames:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]
    pairs: list[tuple[str, str, str]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 5:
            continue
        eng, kor = row[3], row[4]
        if not eng or not kor:
            continue
        eng_s, kor_s = str(eng).strip(), str(kor).strip()
        if eng_s in ("단어", "No") or kor_s == "뜻":
            continue
        if HAS_LATIN.search(eng_s) and HAS_HANGUL.search(kor_s):
            pairs.append((eng_s, kor_s, SOURCE_XLSX))
    wb.close()
    return pairs


def parse_pdf_pairs(path: Path) -> list[tuple[str, str, str]]:
    import fitz

    doc = fitz.open(path)
    text = "".join(page.get_text() for page in doc)
    doc.close()

    lines = [ln.rstrip() for ln in text.splitlines()]
    pairs: list[tuple[str, str, str]] = []
    i = 0
    while i < len(lines):
        ln = lines[i].strip()
        if not (ln.startswith(MARK) or ln.startswith("≅")):
            i += 1
            continue
        body = ln.lstrip(MARK + "≅").strip()
        eng, kor = "", ""
        if "\t" in body:
            eng, kor = body.split("\t", 1)
        else:
            eng = body
        eng, kor = eng.strip(), kor.strip()
        if eng and not HAS_HANGUL.search(eng) and not kor and i + 1 < len(lines):
            nxt = lines[i + 1].strip()
            if HAS_HANGUL.search(nxt) and not nxt.startswith(MARK) and not nxt.startswith("≅"):
                kor = nxt
                i += 2
            else:
                i += 1
        else:
            i += 1
        if eng and kor and HAS_LATIN.search(eng) and HAS_HANGUL.search(kor):
            pairs.append((eng, kor, SOURCE_PDF))
    return pairs


def clean_english(raw: str) -> str:
    s = raw.strip().replace("\u2019", "'")
    s = LESSON_SUFFIX.sub("", s).strip()
    s = re.sub(r"(?<=[a-zA-Z])\d+강\d+번\s*$", "", s).strip()
    s = re.sub(r"(?<=[a-zA-Z])\d+$", "", s).strip()
    return s


def is_korean_meaning(line: str) -> bool:
    s = line.strip()
    if not s or not HAS_HANGUL.search(s):
        return False
    if HAS_LATIN.search(s) and len(HAS_HANGUL.findall(s)) < 2:
        return False
    return True


def is_english_headword(line: str) -> bool:
    s = clean_english(line)
    if not s or len(s) < 2:
        return False
    if ONLY_NUM.match(s) or JUNK_LINE.match(s):
        return False
    if not HAS_LATIN.search(s) or HAS_HANGUL.search(s):
        return False
    return bool(ENG_LINE.match(s))


def parse_hwp_pairs(paragraphs: list[str]) -> list[tuple[str, str, str]]:
    lines = [p.strip() for p in paragraphs if p and p.strip()]
    pairs: list[tuple[str, str, str]] = []
    i = 0
    while i < len(lines):
        eng = lines[i]
        if is_english_headword(eng):
            eng_clean = clean_english(eng)
            if i + 1 < len(lines) and is_korean_meaning(lines[i + 1]):
                pairs.append((eng_clean, lines[i + 1].strip(), SOURCE_HWP))
                i += 2
                continue
        i += 1
    return pairs


def collect_pairs() -> list[tuple[str, str, str]]:
    pairs: list[tuple[str, str, str]] = []
    seen: set[str] = set()

    if XLSX_PATH.exists():
        pairs.extend(parse_xlsx_pairs(XLSX_PATH))
        seen.update(eng_key(e) for e, _, _ in pairs)
        print(f"XLSX: {len(pairs)} pairs from {XLSX_PATH.name}")
    elif HWP_PATH.exists():
        paragraphs = extract_paragraphs(str(HWP_PATH))
        pairs.extend(parse_hwp_pairs(paragraphs))
        seen.update(eng_key(e) for e, _, _ in pairs)
        print(f"HWP fallback: {len(pairs)} pairs")
    else:
        print("No XLSX or HWP source found.")

    if PDF_PATH.exists():
        pdf_added = 0
        for eng, kor, src in parse_pdf_pairs(PDF_PATH):
            key = eng_key(eng)
            if key in seen:
                continue
            pairs.append((eng, kor, src))
            seen.add(key)
            pdf_added += 1
        print(f"PDF supplement: +{pdf_added} (not in xlsx/hwp)")
    return pairs


def load_index() -> tuple[dict, dict]:
    if OUT.exists():
        payload = json.loads(OUT.read_text(encoding="utf-8"))
        return payload.get("words") or {}, payload.get("meta") or {}
    return {}, {}


def merge_phrases_into_passage_db(
    pairs: list[tuple[str, str, str]],
) -> int:
    from merge_idiom_phrases import add as add_phrase  # noqa: E402

    if PHRASE_JSON.exists():
        payload = json.loads(PHRASE_JSON.read_text(encoding="utf-8"))
        index = {
            p: {"mean": v.get("mean", ""), "sources": list(v.get("sources") or [])}
            for p, v in (payload.get("phrases") or {}).items()
        }
    else:
        index = {}

    added = 0
    for eng, kor, src in pairs:
        if " " not in eng.strip() and "/" not in eng.strip():
            continue
        if add_phrase(index, eng, kor, src):
            added += 1

    phrases = sorted(index.keys(), key=lambda x: (-len(x.split()), -len(x), x))
    payload = {
        "meta": {"count": len(phrases), "match_order": "longest_first"},
        "phrases": {
            p: {"mean": index[p]["mean"], "sources": index[p]["sources"]} for p in phrases
        },
    }
    PHRASE_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "/* auto-generated; do not edit */",
        "window.PASSAGE_PHRASE_MEANINGS=(function(){var m=Object.create(null);",
    ]
    for p in phrases:
        lines.append(f"m[{json.dumps(p)}]={json.dumps(index[p]['mean'])};")
    lines.append("return m;})();")
    lines.append("window.PASSAGE_PHRASE_LIST=" + json.dumps(phrases, ensure_ascii=False) + ";")
    PHRASE_JS.write_text("\n".join(lines), encoding="utf-8")
    return added


def write_lookup_js(words: list[str]) -> None:
    body = (
        "window.MARKET_HEADWORD_SET=(function(){var o=Object.create(null);"
        + "".join(f"o[{json.dumps(w)}]=1;" for w in words)
        + "return o;})();"
    )
    OUT_JS.write_text("/* auto-generated; do not edit */\n" + body, encoding="utf-8")


def main() -> None:
    pairs = collect_pairs()
    if not pairs:
        return

    index, meta = load_index()
    before = len(index)
    phrase_lines = 0

    for eng, _kor, src in pairs:
        if " " in eng or "/" in eng:
            phrase_lines += 1
        for w in tokens_from_phrase(eng):
            add(index, w, src)

    words = sorted(index.keys())
    meta.setdefault("source_files", [])
    for path, tag in (
        (XLSX_PATH, SOURCE_XLSX),
        (PDF_PATH, SOURCE_PDF),
        (HWP_PATH, SOURCE_HWP),
    ):
        if path.exists():
            entry = f"{tag}:{path.name}"
            if entry not in meta["source_files"]:
                meta["source_files"].append(entry)
    meta["count"] = len(words)
    meta["ebs_vocab_pairs"] = len(pairs)
    meta["ebs_vocab_phrase_lines"] = phrase_lines
    meta["ebs_primary"] = SOURCE_XLSX if XLSX_PATH.exists() else SOURCE_HWP

    payload = {"meta": meta, "words": {w: index[w] for w in words}}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_lookup_js(words)
    phrase_added = merge_phrases_into_passage_db(pairs)

    print(f"Total pairs merged: {len(pairs)}")
    print(f"Added {len(words) - before} new headwords ({before} -> {len(words)})")
    print(f"Multi-word lines: {phrase_lines}, new phrases: {phrase_added}")
    print(f"Wrote {OUT}")
    print(f"Wrote {OUT_JS} ({OUT_JS.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
